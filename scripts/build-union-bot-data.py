#!/usr/bin/env python3
import json
import hashlib
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


# Начиная с выгрузки 10–16.08.2026 суммы суперюниона уже приведены
# к рублям в самом Excel. Дополнительная конвертация не применяется.
SUPER_LEAGUE_EXCHANGE_RATES = {}

LEAGUE_JACKPOT_REFUNDS = {
    "PPCUNION": 50,
    "VAULT 13": 70,
    "ONL YSTARS": 70,
    "Rbpoker": 70,
    "QUBE": 60,
    "AQUARIUM": 50,
}

LEAGUE_SERVICE_PERCENT = {
    "СССР": 8,
    "VAULT 13": 6,
    "Off Cheats": 8,
    "Bambuk": 6,
    "AQUARIUM": 6,
}

CLUB_SERVICE_PERCENT = {
    "Kings KO": 8,
    "Fish Hunter": 15,
    "Лудоманы": 15,
    "FEBOS": 30,
    "Joker♦️Poker": 8,
    "Joker♦️VIP♦️Poker": 8,
    "Pattaya": 8,
    "Kampashka 21": 8,
    "TipTop": 8,
    "ШАНС": 10,
    "CHICAGO.21": 10,
    "SalamBro": 10,
    "new balance": 10,
    "GKpoker": 10,
    "Амиго": 15,
    "MAGILAN": 10,
    "Collaboration Club": 20,
    "IMMORTALS": 15,
    "Спарта": 15,
    "Beer and Bear": 15,
    "GARAGE": 8,
    "GoRiLaZzz": 10,
    "PC Arena": 8,
    "RealPokerGame": 10,
    "River21": 10,
    "Sibir 70": 10,
    "Два Туза": 8,
    "РИВЕР КЛУБ": 20,
    "T O T": 8,
    "Храм": 10,
}

CLUB_SALARY = {
    "Два Туза": -1500,
    "Kampashka 21": -1500,
    "Joker♦️Poker": -1500,
    "Kings KO": 4500,
}


def write_json(path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def add_value(target, key, value):
    if isinstance(value, (int, float)) and value:
        target[key] = target.get(key, 0) + float(value)


def top(rows, field, reverse=True, count=10):
    return sorted(rows, key=lambda row: row[field], reverse=reverse)[:count]


def league_slug(value):
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug or f"league-{hashlib.sha1(str(value).encode('utf-8')).hexdigest()[:10]}"


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: build-union-bot-data.py INPUT.xlsx OUTPUT_DIR")
    source = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source, read_only=True, data_only=True)

    union = workbook["Union Data"]
    meta = str(union.cell(2, 1).value or "")
    period = re.search(r"Range:\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", meta)
    if not period:
        raise ValueError("Report period was not found in Union Data!A2")
    start_date, end_date = period.groups()
    base = {"startDate": start_date, "endDate": end_date}

    union_headers = [cell.value for cell in next(union.iter_rows(min_row=4, max_row=4))]
    club_metrics = {}
    for row in union.iter_rows(min_row=5, values_only=True):
        if not isinstance(row[1], (int, float)):
            continue
        values = dict(zip(union_headers, row))
        club_id = str(int(row[1]))
        club_metrics[club_id] = {
            "id": club_id,
            "name": str(row[0]),
            "profits": values.get("Profits") or 0,
            "winnings": values.get("Winnings") or 0,
            "rake": values.get("Total Fee") or 0,
            "cashRake": values.get("Ring Game Fee") or 0,
            "mttRake": values.get("MTT Fee") or 0,
            "sngRake": values.get("SNG Fee") or 0,
            "insurance": values.get("Insurance") or 0,
            "regularJackpotFee": values.get("Jackpot Fee") or 0,
            "regularJackpotPayout": values.get("Jackpot Payout") or 0,
            "jackpot21Fee": values.get("Jackpot Fee 21") or 0,
            "jackpot21Payout": values.get("Jackpot Payout 21") or 0,
            "jackpotMttFee": values.get("Jackpot Fee Mtt") or 0,
            "jackpotMttPayout": values.get("Jackpot Payout Mtt") or 0,
            "jackpotTopup": values.get("Jackpot Topup") or 0,
            "games": values.get("Games") or 0,
            "hands": values.get("Hands") or 0,
        }

    members = workbook["Union Member Statistics"]
    member_headers = [cell.value for cell in next(members.iter_rows(min_row=5, max_row=5))]
    players = {}
    club_players = defaultdict(dict)
    member_club_ids = set()
    active_ids = defaultdict(set)
    member_hands = defaultdict(float)
    for row in members.iter_rows(min_row=6, values_only=True):
        if not isinstance(row[0], (int, float)) or not isinstance(row[2], (int, float)):
            continue
        club_id = str(int(row[0]))
        player_id = str(int(row[2]))
        member_club_ids.add(club_id)
        nick = str(row[4] or "").strip() or player_id
        player = players.setdefault(player_id, {
            "id": player_id, "nick": nick, "aliases": set(), "clubs": set(), "agents": set(),
            "winnings": 0.0, "cashRake": 0.0, "mttRake": 0.0, "sngRake": 0.0,
            "insurance": 0.0, "jackpotFee": 0.0, "jackpotPayout": 0.0, "hands": 0.0,
            "winGames": {}, "rakeGames": {},
        })
        player["nick"] = nick
        player["aliases"].add(nick)
        player["clubs"].add(club_id)
        agent_name = str(row[6] or "").strip()
        agent_id = str(int(row[7])) if isinstance(row[7], (int, float)) else ""
        if agent_name or agent_id:
            player["agents"].add(f"{agent_name} ({agent_id})" if agent_name and agent_id else agent_name or agent_id)
        player["winnings"] += float(row[9] or 0)
        player["cashRake"] += float(row[30] or 0)
        player["mttRake"] += float(row[44] or 0)
        player["sngRake"] += float(row[49] or 0)
        player["insurance"] += float(row[53] or 0)
        player["jackpotFee"] += float(row[59] or 0)
        player["jackpotPayout"] += float(row[60] or 0)
        player["hands"] += float(row[61] or 0)
        member_hands[club_id] += float(row[61] or 0)
        if any(float(row[index] or 0) != 0 for index in [9, 30, 44, 49, 53, 59, 60, 61]):
            active_ids[club_id].add(player_id)
        for index in range(10, 30):
            add_value(player["winGames"], str(member_headers[index]), row[index])
        for index in list(range(31, 44)) + list(range(45, 49)) + list(range(50, 53)):
            add_value(player["rakeGames"], str(member_headers[index]), row[index])
        club_player = club_players[club_id].setdefault(player_id, {
            "id": player_id, "nick": nick, "winnings": 0.0, "rake": 0.0,
            "cashRake": 0.0, "mttRake": 0.0, "sngRake": 0.0, "insurance": 0.0,
            "winGames": {}, "rakeGames": {},
        })
        club_player["nick"] = nick
        club_player["winnings"] += float(row[9] or 0)
        club_player["cashRake"] += float(row[30] or 0)
        club_player["mttRake"] += float(row[44] or 0)
        club_player["sngRake"] += float(row[49] or 0)
        club_player["insurance"] += float(row[53] or 0)
        club_player["rake"] = club_player["cashRake"] + club_player["mttRake"] + club_player["sngRake"]
        for index in range(10, 30):
            add_value(club_player["winGames"], str(member_headers[index]), row[index])
        for index in list(range(31, 44)) + list(range(45, 49)) + list(range(50, 53)):
            add_value(club_player["rakeGames"], str(member_headers[index]), row[index])

    for club_id, club in club_metrics.items():
        rows = list(club_players.get(club_id, {}).values())
        for row in rows:
            row["winGames"] = sorted(row["winGames"].items(), key=lambda item: -abs(item[1]))
            row["rakeGames"] = sorted(row["rakeGames"].items(), key=lambda item: -item[1])
        club["jackpotFee"] = club.pop("regularJackpotFee") + club.pop("jackpot21Fee") + club["jackpotMttFee"] + club["jackpotTopup"]
        club["jackpotPayout"] = club.pop("regularJackpotPayout") + club.pop("jackpot21Payout") + club["jackpotMttPayout"]
        club["players"] = len(rows)
        club["topRake"] = top(rows, "rake", True, 5)
        club["topPlus"] = top([row for row in rows if row["winnings"] > 0], "winnings", True, 5)
        club["topMinus"] = top([row for row in rows if row["winnings"] < 0], "winnings", False, 5)
        club["playerRows"] = sorted(rows, key=lambda row: row["nick"].casefold())

    player_rows = []
    for player in players.values():
        player["aliases"] = sorted(player["aliases"])
        player["clubs"] = sorted(player["clubs"])
        player["agents"] = sorted(player["agents"])
        player["rake"] = player["cashRake"] + player["mttRake"] + player["sngRake"]
        player["winGames"] = sorted(player["winGames"].items(), key=lambda item: -abs(item[1]))
        player["rakeGames"] = sorted(player["rakeGames"].items(), key=lambda item: -item[1])
        player_rows.append(player)
    directory_data = {
        **base,
        "clubs": sorted(club_metrics.values(), key=lambda club: club["name"].casefold()),
        "players": sorted(player_rows, key=lambda player: int(player["id"])),
    }
    write_json(output_dir / "union-directory.json", directory_data)

    rake_clubs = [club_metrics[club_id] for club_id in member_club_ids if club_id in club_metrics]
    member_rake_data = {
        **base,
        "clubs": [{"club": club["name"], "clubId": club["id"], "rake": club["rake"]} for club in sorted(rake_clubs, key=lambda club: (-club["rake"], club["name"].casefold()))],
    }
    write_json(output_dir / "union-member-rake-summary.json", member_rake_data)

    games_sheet = workbook["Union Game Statistics"]
    game_headers = [cell.value for cell in next(games_sheet.iter_rows(min_row=5, max_row=5))]
    type_index = game_headers.index("Game Type")
    name_index = game_headers.index("Game Name")
    fee_index = game_headers.index("Fee")
    overlay_index = game_headers.index("Overlay")
    game_rake = defaultdict(float)
    overlays = []
    for row in games_sheet.iter_rows(min_row=6, values_only=True):
        if row[type_index] is not None and isinstance(row[fee_index], (int, float)):
            game_rake[str(row[type_index])] += float(row[fee_index])
        if isinstance(row[overlay_index], (int, float)) and row[overlay_index] != 0:
            overlays.append({"name": str(row[name_index] or "Без названия").strip(), "overlay": float(row[overlay_index])})
    game_rake_data = {**base, "games": [{"name": name, "rake": value} for name, value in sorted(game_rake.items(), key=lambda item: -item[1])]}
    overlay_data = {**base, "tournaments": sorted(overlays, key=lambda row: -row["overlay"])}
    write_json(output_dir / "union-game-rake-summary.json", game_rake_data)
    write_json(output_dir / "union-overlay-summary.json", overlay_data)

    raw_union_rows = []
    for row in union.iter_rows(min_row=5, values_only=True):
        if isinstance(row[1], (int, float)):
            raw_union_rows.append(dict(zip(union_headers, row)))
    super_league_sheet = workbook["Supper Union League Statistics"]
    super_league_headers = [cell.value for cell in next(super_league_sheet.iter_rows(min_row=4, max_row=4))]
    super_league_name_index = super_league_headers.index("      League      ")
    super_league_fee_index = super_league_headers.index("Jackpot Fee Total(Super Union)")
    super_league_payout_index = super_league_headers.index("Jackpot Payout Total(Super Union)")
    super_league_result_index = super_league_headers.index("Total(Super Union)")
    super_league_total_fee_index = super_league_headers.index("FeeTotal(Super Union)")
    super_league_insurance_index = super_league_headers.index("Insurance Total(Super Union)")
    super_league_mtt_fee_index = super_league_headers.index("Jackpot Mtt Fee Total(Super Union)")
    super_league_mtt_payout_index = super_league_headers.index("Jackpot Mtt Payout Total(Super Union)")
    super_league_club_index = super_league_headers.index("      Club      ")
    super_league_player_index = super_league_headers.index("Player ID")
    super_league_insurance_player_index = super_league_headers.index("Insurance")
    super_league_jackpot_fee_player_index = super_league_headers.index("Jackpot Fee")
    super_league_jackpot_payout_player_index = super_league_headers.index("Jackpot Payout")
    jackpot_leagues_by_id = {}
    league_players = defaultdict(dict)
    league_clubs = defaultdict(dict)
    super_league_total_row = None
    for row in super_league_sheet.iter_rows(min_row=5, values_only=True):
        league_label = str(row[super_league_name_index] or "").strip()
        if league_label.casefold() == "total":
            super_league_total_row = row
            continue
        league_match = re.match(r"^(.*)\((\d+)\)$", league_label)
        league_name = league_match.group(1).strip() if league_match else league_label
        league_id = league_match.group(2) if league_match else ""
        exchange_rate = SUPER_LEAGUE_EXCHANGE_RATES.get(league_name, 1)
        if isinstance(row[super_league_player_index], (int, float)) and league_id:
            player_id = str(int(row[super_league_player_index]))
            player = league_players[league_id].setdefault(player_id, {
                "playerId": player_id,
                "nick": str(row[2] or player_id).strip(),
                "clubs": set(),
                "rake": 0.0,
                "winnings": 0.0,
                "insurance": 0.0,
                "jackpotFee": 0.0,
                "jackpotPayout": 0.0,
            })
            club_label = str(row[super_league_club_index] or "").strip()
            club_match = re.match(r"^(.*)\((\d+)\)$", club_label)
            if club_label:
                player["clubs"].add(club_match.group(1).strip() if club_match else club_label)
            player["rake"] += float(row[5] or 0) * exchange_rate
            player["winnings"] += float(row[4] or 0) * exchange_rate
            player["insurance"] += float(row[super_league_insurance_player_index] or 0) * exchange_rate
            player["jackpotFee"] += float(row[super_league_jackpot_fee_player_index] or 0) * exchange_rate
            player["jackpotPayout"] += float(row[super_league_jackpot_payout_player_index] or 0) * exchange_rate
            if club_label:
                club_id = club_match.group(2) if club_match else ""
                club_name = club_match.group(1).strip() if club_match else club_label
                club_key = club_id or club_name.casefold()
                club = league_clubs[league_id].setdefault(club_key, {
                    "clubId": club_id, "club": club_name, "rake": 0.0, "winLose": 0.0,
                    "playerIds": set(), "activePlayerIds": set(),
                })
                club["rake"] += float(row[5] or 0) * exchange_rate
                club["winLose"] += float(row[4] or 0) * exchange_rate
                club["playerIds"].add(player_id)
                if float(row[4] or 0) != 0 or float(row[5] or 0) != 0:
                    club["activePlayerIds"].add(player_id)
        if not isinstance(row[super_league_fee_index], (int, float)) and not isinstance(row[super_league_payout_index], (int, float)):
            continue
        candidate = {
            "league": league_name,
            "leagueId": league_id,
            "fee": float(row[super_league_fee_index] or 0),
            "payout": float(row[super_league_payout_index] or 0),
            "feeTotal": float(row[super_league_total_fee_index] or 0),
            "insurance": float(row[super_league_insurance_index] or 0),
            "mttFee": float(row[super_league_mtt_fee_index] or 0),
            "mttPayout": float(row[super_league_mtt_payout_index] or 0),
            "winLose": float(row[super_league_result_index] or 0),
            "exchangeRate": exchange_rate,
        }
        league_key = league_id or league_name.casefold()
        current = jackpot_leagues_by_id.get(league_key)
        candidate_score = sum(abs(candidate[field]) for field in ("fee", "payout", "feeTotal", "insurance", "mttFee", "mttPayout", "winLose"))
        current_score = sum(abs(current[field]) for field in ("fee", "payout", "feeTotal", "insurance", "mttFee", "mttPayout", "winLose")) if current else -1
        if candidate_score > current_score:
            jackpot_leagues_by_id[league_key] = candidate
    jackpot_leagues = list(jackpot_leagues_by_id.values())
    league_names = {row["leagueId"]: row["league"] for row in jackpot_leagues}
    league_player_rows = []
    for league_id, players_by_id in league_players.items():
        players_for_league = list(players_by_id.values())
        def player_top(field, reverse=True, positive=None):
            selected = players_for_league
            if positive is True:
                selected = [row for row in selected if row[field] > 0]
            elif positive is False:
                selected = [row for row in selected if row[field] < 0]
            selected = sorted(selected, key=lambda row: row[field], reverse=reverse)[:10]
            return [{
                "playerId": row["playerId"],
                "nick": row["nick"],
                "clubs": sorted(row["clubs"]),
                "value": round(row[field], 2),
            } for row in selected]
        league_player_rows.append({
            "leagueId": league_id,
            "league": league_names.get(league_id, league_id),
            "players": sorted([{
                "playerId": row["playerId"],
                "nick": row["nick"],
                "clubs": sorted(row["clubs"]),
                "rake": round(row["rake"], 2),
                "winnings": round(row["winnings"], 2),
                "insurance": round(row["insurance"], 2),
                "jackpotFee": round(row["jackpotFee"], 2),
                "jackpotPayout": round(row["jackpotPayout"], 2),
            } for row in players_for_league], key=lambda row: row["nick"].casefold()),
            "clubs": sorted([
                {
                    "clubId": row["clubId"], "club": row["club"],
                    "rake": round(row["rake"], 2), "winLose": round(row["winLose"], 2),
                    "players": len(row["playerIds"]), "activePlayers": len(row["activePlayerIds"]),
                } for row in league_clubs.get(league_id, {}).values()
            ], key=lambda row: (-row["rake"], row["club"].casefold())),
            "rake": player_top("rake"),
            "minus": player_top("winnings", reverse=False, positive=False),
            "plus": player_top("winnings", positive=True),
        })
    league_player_tops_data = {**base, "leagues": sorted(league_player_rows, key=lambda row: row["league"].casefold())}
    write_json(output_dir / "union-league-player-tops.json", league_player_tops_data)
    super_league_fee = round(sum(row["fee"] * row["exchangeRate"] for row in jackpot_leagues), 4)
    super_league_payout = round(sum(row["payout"] * row["exchangeRate"] for row in jackpot_leagues), 2)
    local_regular_fee = round(sum(float(row.get("Jackpot Fee") or 0) for row in raw_union_rows), 2)
    local_regular_payout = round(sum(float(row.get("Jackpot Payout") or 0) for row in raw_union_rows), 2)
    jackpot_21_fee = round(sum(float(row.get("Jackpot Fee 21") or 0) for row in raw_union_rows), 2)
    jackpot_21_payout = round(sum(float(row.get("Jackpot Payout 21") or 0) for row in raw_union_rows), 2)
    jackpot_mtt_fee = round(sum(row["mttFee"] * row["exchangeRate"] for row in jackpot_leagues), 2)
    jackpot_mtt_payout = round(sum(row["mttPayout"] * row["exchangeRate"] for row in jackpot_leagues), 2)
    jackpot_topup = round(sum(float(row.get("Jackpot Topup") or 0) for row in raw_union_rows), 2)
    jackpot_data = {
        **base,
        "regularFee": local_regular_fee,
        "regularPayout": local_regular_payout,
        "jackpot21Fee": jackpot_21_fee,
        "jackpot21Payout": jackpot_21_payout,
        "jackpotMttFee": jackpot_mtt_fee,
        "jackpotMttPayout": jackpot_mtt_payout,
        "jackpotTopup": jackpot_topup,
        "unclassifiedFee": round(super_league_fee - local_regular_fee - jackpot_21_fee, 4),
        "unclassifiedPayout": round(super_league_payout - local_regular_payout - jackpot_21_payout, 2),
        "totalFee": round(super_league_fee + jackpot_mtt_fee + jackpot_topup, 4),
        "totalPayout": round(super_league_payout + jackpot_mtt_payout, 2),
        "calculations": {
            "winLose": float(super_league_total_row[super_league_result_index] or 0) if super_league_total_row else 0,
            "fee": round(sum(row["feeTotal"] * row["exchangeRate"] for row in jackpot_leagues), 4),
            "insurance": round(sum(row["insurance"] * row["exchangeRate"] for row in jackpot_leagues), 2),
            "overlay": round(sum(float(row["overlay"] or 0) for row in overlays), 2),
        },
        "leagues": sorted(jackpot_leagues, key=lambda row: (-row["fee"], row["league"].casefold())),
    }
    write_json(output_dir / "union-jackpot-summary.json", jackpot_data)

    league_reports = []
    for row in jackpot_leagues:
        if row["league"].casefold() == "анти-рег":
            continue
        exchange_rate = row["exchangeRate"]
        winnings = round(row["winLose"], 2)
        commission = round(row["feeTotal"] * exchange_rate, 2)
        balance = round(winnings + commission, 2)
        fraud = 0
        overly = 0
        balance_final = round(balance + fraud + overly, 2)
        promo = 0
        service_percent = LEAGUE_SERVICE_PERCENT.get(row["league"], 5)
        service = round(-commission * service_percent / 100, 2)
        refund_percent = LEAGUE_JACKPOT_REFUNDS.get(row["league"], 0)
        # Откат союзу считается со всего его джекпот-сбора: кеш + Jackpot 21 + MTT.
        jackpot_refund = int((row["fee"] + row["mttFee"]) * exchange_rate * refund_percent / 100)
        total = round(balance_final + promo + service + jackpot_refund, 2)
        league_reports.append({
            "league": row["league"],
            "leagueId": row["leagueId"],
            "startDate": start_date,
            "endDate": end_date,
            "imagePath": f"/assets/reports/unions/{start_date}_{end_date}/{league_slug(row['league'])}.png",
            "metrics": {
                "winnings": winnings,
                "commission": commission,
                "balance": balance,
                "fraud": fraud,
                "overly": overly,
                "balanceFinal": balance_final,
                "promo": promo,
                "servicePercent": service_percent,
                "service": service,
                "jackpotRefundPercent": refund_percent,
                "jackpotRefund": jackpot_refund,
                "total": total,
            },
        })
    league_report_data = {**base, "reports": league_reports}
    write_json(output_dir / "union-league-reports.json", league_report_data)

    group_sheet = workbook["Supper Union Group Statistics"]
    group_headers = [cell.value for cell in next(group_sheet.iter_rows(min_row=4, max_row=4))]
    group_club_index = group_headers.index("      Club      ")
    group_winnings_index = group_headers.index("Total")
    group_fee_index = group_headers.index("FeeTotal")
    group_club_metrics = {}
    for row in group_sheet.iter_rows(min_row=5, values_only=True):
        club_match = re.match(r"^(.*)\((\d+)\)$", str(row[group_club_index] or "").strip())
        if not club_match:
            continue
        club_name, club_id = club_match.group(1).strip(), club_match.group(2)
        candidate = {
            "id": club_id,
            "name": club_name,
            "winnings": float(row[group_winnings_index] or 0),
            "rake": float(row[group_fee_index] or 0),
        }
        current = group_club_metrics.get(club_id)
        if not current or abs(candidate["winnings"]) + abs(candidate["rake"]) > abs(current["winnings"]) + abs(current["rake"]):
            group_club_metrics[club_id] = candidate

    anti_reg_clubs = {club_id: (row["name"], "184691") for club_id, row in group_club_metrics.items()}

    club_reports = []
    for club_id, (club_name, league_id) in anti_reg_clubs.items():
        source_metrics = club_metrics.get(club_id) or group_club_metrics.get(club_id)
        if not source_metrics:
            continue
        league_name = "Анти-Рег"
        winnings = round(source_metrics["winnings"], 2)
        commission = round(source_metrics["rake"], 2)
        balance = round(winnings + commission, 2)
        service_percent = CLUB_SERVICE_PERCENT.get(club_name, 10)
        service = round(-commission * service_percent / 100, 2)
        salary = CLUB_SALARY.get(club_name, 0)
        total = round(balance + salary + service, 2)
        club_reports.append({
            "club": club_name,
            "clubId": club_id,
            "league": league_name,
            "leagueId": league_id,
            "startDate": start_date,
            "endDate": end_date,
            "imagePath": f"/assets/reports/clubs/{start_date}_{end_date}/{league_slug(club_name)}-{club_id}.png",
            "metrics": {
                "winnings": winnings,
                "commission": commission,
                "balance": balance,
                "fraud": 0,
                "overly": 0,
                "balanceFinal": balance,
                "promo": 0,
                "salary": salary,
                "servicePercent": service_percent,
                "service": service,
                "jackpotRefundPercent": 0,
                "jackpotRefund": 0,
                "total": total,
            },
        })
    club_reports.sort(key=lambda row: (row["league"].casefold(), row["club"].casefold()))
    club_report_data = {**base, "reports": club_reports}
    write_json(output_dir / "union-club-reports.json", club_report_data)

    player_tops_data = {
        **base,
        "rake": [{"playerId": row["id"], "nick": row["nick"], "clubs": [club_metrics[cid]["name"] for cid in row["clubs"] if cid in club_metrics], "value": row["rake"]} for row in top(player_rows, "rake")],
        "minus": [{"playerId": row["id"], "nick": row["nick"], "clubs": [club_metrics[cid]["name"] for cid in row["clubs"] if cid in club_metrics], "value": row["winnings"]} for row in top([row for row in player_rows if row["winnings"] < 0], "winnings", False)],
        "plus": [{"playerId": row["id"], "nick": row["nick"], "clubs": [club_metrics[cid]["name"] for cid in row["clubs"] if cid in club_metrics], "value": row["winnings"]} for row in top([row for row in player_rows if row["winnings"] > 0], "winnings")],
    }
    write_json(output_dir / "union-player-tops.json", player_tops_data)

    activity_rows = []
    for club_id, club in club_metrics.items():
        row = {
            "club": club["name"], "clubId": club_id, "rake": club["rake"], "games": club["games"],
            "hands": member_hands[club_id], "activePlayers": len(active_ids[club_id]),
        }
        if row["rake"] or row["games"] or row["hands"] or row["activePlayers"]:
            row["rakePerPlayer"] = round(row["rake"] / row["activePlayers"], 2) if row["activePlayers"] else 0
            activity_rows.append(row)
    activity_data = {
        **base,
        "activeClubs": len(activity_rows),
        "activePlayers": sum(row["activePlayers"] for row in activity_rows),
        "games": sum(row["games"] for row in activity_rows),
        "hands": sum(row["hands"] for row in activity_rows),
        "topPlayers": top(activity_rows, "activePlayers"),
        "topGames": top(activity_rows, "games"),
        "topHands": top(activity_rows, "hands"),
        "topRakePerPlayer": top(activity_rows, "rakePerPlayer"),
    }
    write_json(output_dir / "union-activity-summary.json", activity_data)

    archive_path = output_dir / "union-periods.json"
    archive = json.loads(archive_path.read_text(encoding="utf-8")) if archive_path.exists() else {"periods": []}
    bundle = {
        **base,
        "directory": directory_data,
        "memberRake": member_rake_data,
        "games": game_rake_data,
        "overlays": overlay_data,
        "jackpot": jackpot_data,
        "leagueReports": league_report_data,
        "clubReports": club_report_data,
        "playerTops": player_tops_data,
        "leaguePlayerTops": league_player_tops_data,
        "activity": activity_data,
    }
    periods = [row for row in archive.get("periods", []) if not (row.get("startDate") == start_date and row.get("endDate") == end_date)]
    periods.append(bundle)
    archive["periods"] = sorted(periods, key=lambda row: row["endDate"], reverse=True)
    write_json(archive_path, archive)
    print(json.dumps({"period": [start_date, end_date], "clubs": len(club_metrics), "memberClubs": len(member_club_ids), "players": len(player_rows), "overlays": len(overlays)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
