#!/usr/bin/env python3
"""Extract club-level rows from a Poker21 Plus workbook for the XLSX builder."""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract-club-report-data.py INPUT.xlsx OUTPUT.json")
    source, target = Path(sys.argv[1]), Path(sys.argv[2])
    workbook = load_workbook(source, read_only=True, data_only=True)

    union = workbook["Union Data"]
    meta = str(union.cell(2, 1).value or "")
    match = re.search(r"Range:\s*(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", meta)
    if not match:
        raise ValueError("Report period was not found")
    start_date, end_date = match.groups()
    union_headers = [clean(cell.value) for cell in next(union.iter_rows(min_row=4, max_row=4))]
    metrics = {}
    for row in union.iter_rows(min_row=5, values_only=True):
        if isinstance(row[1], (int, float)):
            metrics[str(int(row[1]))] = dict(zip(union_headers, map(clean, row)))

    members = workbook["Union Member Statistics"]
    headers = [clean(cell.value) or f"Column {index}" for index, cell in enumerate(next(members.iter_rows(min_row=5, max_row=5)), 1)]
    rows_by_club = defaultdict(list)
    names = {}
    for row in members.iter_rows(min_row=6, values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        club_id = str(int(row[0]))
        names[club_id] = str(row[1] or metrics.get(club_id, {}).get("Club Name") or club_id)
        rows_by_club[club_id].append([clean(value) for value in row])

    payload = {
        "source": source.name,
        "startDate": start_date,
        "endDate": end_date,
        "headers": headers,
        "clubs": [
            {"id": club_id, "name": names[club_id], "metrics": metrics.get(club_id, {}), "rows": rows}
            for club_id, rows in sorted(rows_by_club.items(), key=lambda item: names[item[0]].casefold())
        ],
    }
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"period": [start_date, end_date], "clubs": len(payload["clubs"]), "rows": sum(len(c["rows"]) for c in payload["clubs"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
